from __future__ import annotations

import csv
import json
import os
import uuid
from typing import Any

from fastapi import APIRouter, UploadFile, File, Depends, HTTPException
from pydantic import BaseModel
from sqlalchemy.orm import Session

from openpyxl import load_workbook

from ..deps import get_db, get_current_user
from ..models import AudienceFile, User
from ..settings import settings

router = APIRouter(prefix="/api/audience", tags=["audience"])


def _ensure_dir(path: str) -> None:
    os.makedirs(path, exist_ok=True)


def _looks_like_email(s: str) -> bool:
    s = (s or "").strip()
    return ("@" in s) and ("." in s) and (len(s) <= 254)


def _read_csv_columns_and_rows(path: str, max_rows: int = 2000) -> tuple[list[str], list[dict[str, Any]]]:
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        cols = reader.fieldnames or []
        rows: list[dict[str, Any]] = []
        for i, row in enumerate(reader):
            if i >= max_rows:
                break
            rows.append(row)
        return cols, rows


def _read_xlsx_columns_and_rows(path: str, max_rows: int = 2000) -> tuple[list[str], list[dict[str, Any]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active  # ilk sheet

    rows_iter = ws.iter_rows(values_only=True)

    try:
        header = next(rows_iter)
    except StopIteration:
        return [], []

    cols = [str(c).strip() if c is not None else "" for c in header]
    cols = [c if c else f"column_{i+1}" for i, c in enumerate(cols)]

    out_rows: list[dict[str, Any]] = []
    for i, r in enumerate(rows_iter):
        if i >= max_rows:
            break
        row_dict = {}
        for j, col in enumerate(cols):
            val = r[j] if j < len(r) else None
            row_dict[col] = "" if val is None else str(val).strip()
        out_rows.append(row_dict)

    return cols, out_rows


def _read_columns_and_rows(path: str, ext: str, max_rows: int = 2000) -> tuple[list[str], list[dict[str, Any]]]:
    if ext == ".csv":
        return _read_csv_columns_and_rows(path, max_rows=max_rows)
    if ext == ".xlsx":
        return _read_xlsx_columns_and_rows(path, max_rows=max_rows)
    raise HTTPException(status_code=400, detail="Unsupported file type")


class ValidateIn(BaseModel):
    file_id: int
    email_column: str


@router.post("/upload")
async def upload_audience(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    _ensure_dir(os.path.join(settings.STORAGE_DIR, "audience"))

    ext = (os.path.splitext(file.filename or "")[1] or "").lower()
    if ext not in [".csv", ".xlsx"]:
        raise HTTPException(status_code=400, detail="Only .csv or .xlsx supported")

    storage_name = f"{uuid.uuid4().hex}{ext}"
    storage_path = os.path.join(settings.STORAGE_DIR, "audience", storage_name)

    contents = await file.read()
    with open(storage_path, "wb") as out:
        out.write(contents)

    cols, _rows = _read_columns_and_rows(storage_path, ext=ext, max_rows=5)

    r = AudienceFile(
        user_id=user.id,
        original_name=file.filename or "audience",
        storage_path=storage_path,
        columns_json=json.dumps({"columns": cols, "ext": ext}, ensure_ascii=False),
    )
    db.add(r)
    db.commit()
    db.refresh(r)

    return {
        "ok": True,
        "id": r.id,
        "original_name": r.original_name,
        "columns": cols,
        "created_at": r.created_at.isoformat(),
    }


@router.get("")
def list_audience(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    rows = (
        db.query(AudienceFile)
        .filter(AudienceFile.user_id == user.id)
        .order_by(AudienceFile.id.desc())
        .all()
    )
    out = []
    for r in rows:
        try:
            cols = json.loads(r.columns_json or "{}").get("columns", [])
        except Exception:
            cols = []
        out.append({
            "id": r.id,
            "original_name": r.original_name,
            "columns": cols,
            "created_at": r.created_at.isoformat(),
        })
    return out


@router.delete("/{file_id}")
def delete_audience(file_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    r = db.query(AudienceFile).filter(AudienceFile.id == file_id, AudienceFile.user_id == user.id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Not found")

    try:
        if r.storage_path and os.path.exists(r.storage_path):
            os.remove(r.storage_path)
    except Exception:
        pass

    db.delete(r)
    db.commit()
    return {"ok": True}


@router.post("/validate")
def validate_audience(payload: ValidateIn, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    r = db.query(AudienceFile).filter(AudienceFile.id == payload.file_id, AudienceFile.user_id == user.id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Not found")

    if not r.storage_path or not os.path.exists(r.storage_path):
        raise HTTPException(status_code=400, detail="File missing on disk")

    ext = (os.path.splitext(r.storage_path)[1] or "").lower()

    cols, rows = _read_columns_and_rows(r.storage_path, ext=ext, max_rows=2000)
    if payload.email_column not in cols:
        raise HTTPException(status_code=400, detail=f"Column not found: {payload.email_column}")

    total = 0
    valid = 0
    invalid_samples = []

    for row in rows:
        total += 1
        email = (row.get(payload.email_column) or "").strip()
        if _looks_like_email(email):
            valid += 1
        else:
            if len(invalid_samples) < 10:
                invalid_samples.append(email)

    # preview (opsiyonel)
    p_cols, p_rows = _read_columns_and_rows(r.storage_path, ext=ext, max_rows=5)

    return {
        "ok": True,
        "total": total,
        "valid": valid,
        "invalid": total - valid,
        "invalid_samples": invalid_samples,
        "preview": p_rows,     # opsiyonel
        "preview_rows": len(p_rows),
        "columns": p_cols,
    }
