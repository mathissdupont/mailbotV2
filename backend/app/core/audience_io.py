from __future__ import annotations

import csv
import os
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import load_workbook


def looks_like_email(s: str) -> bool:
    s = (s or "").strip()
    return ("@" in s) and ("." in s) and (len(s) <= 254)


def read_csv_columns_and_rows(path: str, max_rows: int = 200000) -> Tuple[List[str], List[Dict[str, Any]]]:
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        cols = reader.fieldnames or []
        rows: List[Dict[str, Any]] = []
        for i, row in enumerate(reader):
            if i >= max_rows:
                break
            rows.append(row)
        return cols, rows


def read_xlsx_columns_and_rows(path: str, max_rows: int = 200000) -> Tuple[List[str], List[Dict[str, Any]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)

    try:
        header = next(it)
    except StopIteration:
        return [], []

    cols = [str(c).strip() if c is not None else "" for c in header]
    cols = [c if c else f"column_{i+1}" for i, c in enumerate(cols)]

    rows: List[Dict[str, Any]] = []
    for i, r in enumerate(it):
        if i >= max_rows:
            break
        d: Dict[str, Any] = {}
        for j, col in enumerate(cols):
            val = r[j] if j < len(r) else None
            d[col] = "" if val is None else str(val).strip()
        rows.append(d)

    return cols, rows


def read_columns_and_rows(path: str, max_rows: int = 200000) -> Tuple[List[str], List[Dict[str, Any]]]:
    ext = (os.path.splitext(path)[1] or "").lower()
    if ext == ".csv":
        return read_csv_columns_and_rows(path, max_rows=max_rows)
    if ext == ".xlsx":
        return read_xlsx_columns_and_rows(path, max_rows=max_rows)
    raise ValueError(f"Unsupported file type: {ext}")


def iter_emails_from_audience_file(storage_path: str, email_column: str) -> Iterable[str]:
    cols, rows = read_columns_and_rows(storage_path, max_rows=500000)
    if email_column not in cols:
        raise ValueError(f"Email column not found: {email_column}")
    for r in rows:
        e = (r.get(email_column) or "").strip()
        if looks_like_email(e):
            yield e
